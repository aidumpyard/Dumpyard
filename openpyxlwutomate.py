import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill

# Sample data for DF1 and DF2
data1 = {
    'A': [1, 2, 3],
    'B': [4, 5, 6],
    'C': [7, 8, 9],
    'D': [10, 11, 12]
}

data2 = {
    'A': [1, 2, 3],
    'B': [4, 5, 6],
    'C': [10, 11, 12],
    'D': [13, 14, 15]
}

DF1 = pd.DataFrame(data1)
DF2 = pd.DataFrame(data2)

# Create DF_delta with the differences in columns C and D
DF_delta = DF2.copy()
DF_delta['C'] = DF2['C'] - DF1['C']
DF_delta['D'] = DF2['D'] - DF1['D']

# Merge DF1, DF2, and DF_delta on columns A and B
merged_df = pd.merge(DF1, DF2, on=['A', 'B'], suffixes=('_DF1', '_DF2'))
merged_df = pd.merge(merged_df, DF_delta, on=['A', 'B'], suffixes=('', '_delta'))

# Rename columns to keep only 'A', 'B', 'C', 'D'
merged_df.columns = ['A', 'B', 'C', 'D', 'C', 'D', 'C', 'D']

# Create a new Excel writer object
writer = pd.ExcelWriter('output.xlsx', engine='openpyxl')

# Write the merged DataFrame to the Excel file starting from cell C7
merged_df.to_excel(writer, sheet_name='Sheet1', startrow=6, startcol=2, header=False, index=False)

# Access the workbook and the worksheet to customize the header
workbook = writer.book
worksheet = writer.sheets['Sheet1']

# Set the column headers
for col_num, value in enumerate(['A', 'B', 'C', 'D', 'C', 'D', 'C', 'D'], start=3):
    worksheet.cell(row=7, column=col_num).value = value

# Merge cells for the top header
worksheet.merge_cells('C1:H1')
worksheet.cell(row=1, column=3).value = 'DF Analysis'

# Merge cells for the second header
worksheet.merge_cells('C2:E2')
worksheet.cell(row=2, column=3).value = 'DF1'
worksheet.merge_cells('F2:H2')
worksheet.cell(row=2, column=6).value = 'DF2'
worksheet.merge_cells('I2:K2')
worksheet.cell(row=2, column=9).value = 'DF Delta'

# Define the style for the second header
header_font = Font(bold=True, color='FFFFFF')
header_alignment = Alignment(horizontal='center', wrap_text=True)
header_border = Border(bottom=Side(style='thin', color='FFFFFF'))
header_fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')

# Apply the style to the second header
for cell in worksheet['C2:K2'][0]:
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = header_border
    cell.fill = header_fill

# Set the row height for the second header
worksheet.row_dimensions[2].height = 20

# Set the column width for A, B, C, D columns
for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
    worksheet.column_dimensions[col].width = 20

# Save the Excel file
writer.close()