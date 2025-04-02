import pandas as pd
import re
from datetime import datetime
from openpyxl import load_workbook

def get_latest_sheet_name(sheet_names):
    month_year_pattern = r'^(January|February|March|April|May|June|July|August|September|October|November|December) \d{4} IFF$'
    valid_sheets = [name for name in sheet_names if re.match(month_year_pattern, name)]
    
    def extract_date(sheet_name):
        return datetime.strptime(sheet_name.replace(' IFF', ''), "%B %Y")
    
    if not valid_sheets:
        return None
    latest_sheet = max(valid_sheets, key=extract_date)
    return latest_sheet

def find_last_active_row(sheet_df):
    for idx in reversed(range(len(sheet_df))):
        row = sheet_df.iloc[idx]
        if all(pd.notna(row.iloc[:4])) and all(str(cell).strip() != '' for cell in row.iloc[:4]):
            return idx + 1  # return index of next empty row
    return 0

def update_excel_with_df(file_path, df_to_append):
    # Load workbook
    workbook = load_workbook(file_path)
    latest_sheet_name = get_latest_sheet_name(workbook.sheetnames)
    if not latest_sheet_name:
        print("No valid sheet found.")
        return
    
    print(f"Updating sheet: {latest_sheet_name}")
    sheet = workbook[latest_sheet_name]
    
    # Read sheet into pandas DataFrame
    sheet_df = pd.DataFrame(sheet.values)
    
    # Find next empty row based on first 4 columns
    start_row = find_last_active_row(sheet_df)
    
    # Write DataFrame to sheet row by row
    for r_idx, row in df_to_append.iterrows():
        for c_idx, val in enumerate(row):
            sheet.cell(row=start_row + 1 + r_idx, column=c_idx + 1, value=val)
    
    # Save workbook
    workbook.save(file_path)
    print(f"Data written starting at row {start_row + 1}.")

# Example usage:
# df_data = pd.DataFrame([[1, 2, 3, 4], [5, 6, 7, 8]], columns=['A', 'B', 'C', 'D'])
# update_excel_with_df("path_to_file.xlsx", df_data)