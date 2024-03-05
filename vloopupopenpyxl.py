from openpyxl import Workbook

# Create a new workbook and select the active worksheet
wb = Workbook()
ws = wb.active

# Set some sample data for the VLOOKUP
ws['A1'] = 'Apple'
ws['B1'] = 100
ws['A2'] = 'Banana'
ws['B2'] = 150
ws['A3'] = 'Orange'
ws['B3'] = 200

# Set a VLOOKUP formula in cell C1 to find the price of 'Banana'
ws['C1'] = "=VLOOKUP(\"Banana\", A1:B3, 2, FALSE)"

# Save the workbook
wb.save('vlookup_example.xlsx')