import openpyxl
import re
import os

def looks_like_filepath(s):
    if not isinstance(s, str):
        return False
    return (
        ("/" in s or "\\" in s) and 
        os.path.splitext(s)[1].lower() in ['.pdf', '.xlsx', '.xls', '.docx', '.csv', '.txt']
    )

# Load workbook
file_path = "your_excel_file.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# Step 1: Get max APP number and row
max_app_number = 0
max_app_row = None
pattern = re.compile(r'APP(\d{3,})')

for row in range(36, ws.max_row + 1):
    val = ws[f"F{row}"].value
    if isinstance(val, str):
        match = pattern.search(val)
        if match:
            num = int(match.group(1))
            if num > max_app_number:
                max_app_number = num
                max_app_row = row

print(f"Highest APP number found: APP{max_app_number} at row {max_app_row}")

# Step 2: Assign hyperlinks only to actual file paths
next_app_number = max_app_number + 1

for row in range(36, ws.max_row + 1):
    cell = ws[f"F{row}"]
    if looks_like_filepath(cell.value):
        app_name = f"APP{next_app_number}"
        cell.value = f'=HYPERLINK("{cell.value}", "{app_name}")'
        next_app_number += 1

# Save file
wb.save("updated_" + file_path)
print("Done.")