import pandas as pd
import openpyxl

# Assuming merged_df is your final DataFrame

# Create a new Excel writer object
writer = pd.ExcelWriter('output.xlsx', engine='openpyxl')

# Write the merged DataFrame to the Excel file starting from cell C7
merged_df.to_excel(writer, sheet_name='Sheet1', startrow=6, startcol=2, index=False)

# Access the workbook and the worksheet to customize the header
workbook = writer.book
worksheet = writer.sheets['Sheet1']

# Merge cells for the top header
worksheet.merge_cells('C1:H1')
worksheet['C1'] = 'DF Analysis'

# Merge cells for the second header
worksheet.merge_cells('C2:E2')
worksheet['C2'] = 'DF1'
worksheet.merge_cells('F2:H2')
worksheet['F2'] = 'DF2'
worksheet.merge_cells('I2:K2')
worksheet['I2'] = 'DF Delta'

# Save the Excel file
writer.save()