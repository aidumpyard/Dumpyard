import openpyxl

# Load the Excel file
file_path = "your_excel_file.xlsx"  # Change to your actual file
wb = openpyxl.load_workbook(file_path)
ws = wb.active  # Change to specific sheet if needed: wb["SheetName"]

# Iterate through column F from row 36 onwards
for row in range(36, ws.max_row + 1):  
    cell = ws[f"F{row}"]
    if cell.value and isinstance(cell.value, str):  # Ensure the cell contains a string
        file_link = cell.value  # Assuming it contains a file path
        cell.value = f'=HYPERLINK("{file_link}", "Open File")'  # Insert hyperlink

# Save the modified file
wb.save("updated_" + file_path)
print("Hyperlinks inserted successfully!")