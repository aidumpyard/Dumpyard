import pandas as pd
import os
from openpyxl import load_workbook

# Step 1: Load the dataframes
email_dump = pd.read_csv('email_dump.csv')  # Adjust this path as per the file location
main_df = pd.read_csv('main_df.csv')  # Adjust this path as per the file location

# Step 2: Open the existing Excel file
excel_file = 'IMCS.xlsx'
book = load_workbook(excel_file)
sheet = book['Sheet2']  # Assuming the sheet is named 'Sheet2'

# Step 3: Merge the DataFrames based on the given condition
merged_df = pd.merge(email_dump, main_df, left_on=['entity', 'cpty_name'], right_on=['entity', 'cpty_name'], how='inner')

# Step 4: Append the .msg file path in the 'Approval' column
merged_df['Approval'] = merged_df['filepath'].apply(lambda x: os.path.basename(x))  # Insert the file name

# Step 5: Append new rows to the existing Excel sheet
# Open the sheet and find the last row with data
last_row = sheet.max_row

# Iterate through the merged dataframe and append rows to the excel sheet
for row in merged_df.itertuples(index=False):
    sheet.append(row)

# Save the changes to the Excel file
book.save(excel_file)

# Verify if the new rows are added correctly
print("New rows have been appended to the Excel file.")