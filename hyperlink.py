import openpyxl

# Load the Excel file
file_path = "your_excel_file.xlsx"  # Replace with your actual file
wb = openpyxl.load_workbook(file_path)
ws = wb.active  # Change if using a specific sheet: wb["SheetName"]

# Start inserting hyperlinks from row 36 onwards in column F
app_number = 612  # Start from APP612

for row in range(36, ws.max_row + 1):  
    cell = ws[f"F{row}"]
    
    if cell.value and isinstance(cell.value, str):  # Ensure the cell contains text
        file_path = cell.value  # Extract file path from cell
        app_name = f"APP{app_number}"  # Generate APP number sequentially
        cell.value = f'=HYPERLINK("{file_path}", "{app_name}")'  # Insert hyperlink
        app_number += 1  # Increment APP number for the next row

# Save the updated Excel file
wb.save("updated_" + file_path)
print("Hyperlinks with sequential APP numbers added successfully!")