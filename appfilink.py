import openpyxl
import re
import os
import shutil

# Helper: check if a string is a valid file path
def looks_like_filepath(s):
    if not isinstance(s, str):
        return False
    return (
        os.path.isfile(s) and
        os.path.splitext(s)[1].lower() in ['.pdf', '.xlsx', '.xls', '.docx', '.csv', '.txt']
    )

# Load Excel
excel_path = "your_excel_file.xlsx"  # Change to your actual file name
wb = openpyxl.load_workbook(excel_path)
ws = wb.active  # Or use wb["SheetName"]

# Step 1: Find highest APP number in column F (row 36+)
max_app_number = 0
max_app_row = None
pattern = re.compile(r'APP(\d{3,})')

for row in range(36, ws.max_row + 1):
    val = ws[f"F{row}"].value
    if isinstance(val, str):
        match = pattern.search(val)
        if match:
            num = int(match.group(1))
            if num > max_app_number:
                max_app_number = num
                max_app_row = row

print(f"Highest APP number found: APP{max_app_number} at row {max_app_row}")

# Step 2: Process each file path
next_app_number = max_app_number + 1

for row in range(36, ws.max_row + 1):
    cell = ws[f"F{row}"]
    original_path = cell.value

    if looks_like_filepath(original_path):
        dir_path, original_filename = os.path.split(original_path)
        ext = os.path.splitext(original_filename)[1]
        app_name = f"APP{next_app_number}"
        
        # New file name: APP###_original.ext
        new_filename = f"{app_name}_{original_filename}"
        new_path = os.path.join(dir_path, new_filename)

        # Rename the actual file (move to new name)
        try:
            shutil.move(original_path, new_path)
            print(f"Renamed: {original_filename} -> {new_filename}")
        except Exception as e:
            print(f"Failed to rename {original_path}: {e}")
            continue

        # Update cell with hyperlink to new file
        cell.value = f'=HYPERLINK("{new_path}", "{app_name}")'
        next_app_number += 1

# Step 3: Save the Excel file
wb.save("updated_" + excel_path)
print("Excel updated with renamed files and hyperlinks.")