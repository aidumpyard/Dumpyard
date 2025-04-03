import openpyxl
import re

# Load Excel file
file_path = "your_excel_file.xlsx"  # Replace with actual file name
wb = openpyxl.load_workbook(file_path)
ws = wb.active  # Or wb["SheetName"]

# Step 1: Find highest APP number in column F starting from row 36
max_app_number = 0
pattern = re.compile(r'APP(\d{3,})')

for row in range(36, ws.max_row + 1):
    cell_value = ws[f"F{row}"].value
    if cell_value and isinstance(cell_value, str):
        match = pattern.search(cell_value)
        if match:
            num = int(match.group(1))
            if num > max_app_number:
                max_app_number = num

# Step 2: Start assigning from the next APP number
next_app_number = max_app_number + 1

# Step 3: Replace any file paths (that are not already APP hyperlinks) with new hyperlinks
for row in range(36, ws.max_row + 1):
    cell = ws[f"F{row}"]
    if cell.value and isinstance(cell.value, str):
        # Only update if it's a file path (i.e., doesn't already start with =HYPERLINK)
        if not cell.value.strip().startswith("=HYPERLINK"):
            file_path = cell.value
            app_name = f"APP{next_app_number}"
            cell.value = f'=HYPERLINK("{file_path}", "{app_name}")'
            next_app_number += 1

# Save the updated file
wb.save("updated_" + file_path)
print("Hyperlinks created and numbered based on existing APP values.")