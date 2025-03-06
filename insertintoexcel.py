import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image
import os

# Load DataFrames
df1 = pd.DataFrame({
    "cpty": ["A", "B"],
    "ccpty": ["X", "Y"],
    "filepath": ["path/to/file1.png", "path/to/file2.png"],
    "date": ["2024-03-06", "2024-03-07"]
})

df2 = pd.DataFrame({
    "cpty": ["A", "B"],
    "ccpty": ["X", "Y"],
    "value": [100, 200]
})

# Dictionary for replacement
replace_dict = {"A": "Alpha", "B": "Beta", "X": "Xi", "Y": "Ypsilon"}

# Replace values in df1
df1.replace({"cpty": replace_dict, "ccpty": replace_dict}, inplace=True)

# Merge df1 and df2 on 'cpty' and 'ccpty'
merged_df = df1.merge(df2, on=["cpty", "ccpty"], how="inner")

# Load existing Excel file or create new one
excel_path = "output.xlsx"
sheet_name = "Sheet1"

try:
    book = openpyxl.load_workbook(excel_path)
    sheet = book[sheet_name]
except FileNotFoundError:
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = sheet_name
    sheet.append(["Cpty", "CCpty", "Value", "Filename", "File"])  # Header

# Append data to Excel
for index, row in merged_df.iterrows():
    last_row = sheet.max_row + 1  # Get next row index
    
    # Insert data (Cpty, CCpty, Value, Filename)
    sheet.cell(row=last_row, column=1, value=row["cpty"])
    sheet.cell(row=last_row, column=2, value=row["ccpty"])
    sheet.cell(row=last_row, column=3, value=row["value"])
    sheet.cell(row=last_row, column=4, value=row["filepath"])

    # Insert image if file exists
    if os.path.exists(row["filepath"]):
        img = Image(row["filepath"])
        sheet.add_image(img, f"E{last_row}")  # Insert into column E

# Save the updated Excel file
book.save(excel_path)
print(f"Excel file '{excel_path}' updated successfully!")