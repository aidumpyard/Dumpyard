import pandas as pd
import os
from openpyxl import load_workbook

# Load the dataframes
email_dump = pd.read_csv('email_dump.csv')  # Replace with actual file path
main_df = pd.read_csv('main_df.csv')  # Replace with actual file path

# Load the existing Excel file and sheet
excel_file = 'IMCS.xlsx'
wb = load_workbook(excel_file)
ws = wb['Sheet2']  # Replace with actual sheet name if different

# Merge dataframes based on given condition
merged_df = email_dump.merge(
    main_df,
    left_on=['entity', 'cpty_name'],
    right_on=['entity', 'cpty_name'],
    how='inner'
)

# Add the .msg file name to the 'Approval' column
merged_df['Approval'] = merged_df['filepath'].apply(lambda x: os.path.basename(x))

# Get the last row in the existing sheet
last_row = ws.max_row  # Find the last row with data

# Append new rows to the sheet
for r in merged_df.itertuples(index=False, name=None):
    ws.append(r)

# Save changes
wb.save(excel_file)
wb.close()

print("New rows appended successfully.")