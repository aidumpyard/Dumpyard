import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill

class ExcelFormatter:
    def __init__(self, dataframe, start_row, start_col, header_config):
        self.dataframe = dataframe
        self.start_row = start_row
        self.start_col = start_col
        self.header_config = header_config
        self.writer = pd.ExcelWriter('output.xlsx', engine='openpyxl')
        self.workbook = self.writer.book
        self.worksheet = self.writer.sheets['Sheet1']

    def write_data(self):
        self.dataframe.to_excel(self.writer, sheet_name='Sheet1', startrow=self.start_row - 1, startcol=self.start_col - 1, header=False, index=False)

    def format_headers(self):
        row = self.start_row
        for layer in self.header_config:
            col = self.start_col
            for header in layer:
                cell = self.worksheet.cell(row=row, column=col)
                cell.value = header['label']
                if 'merge' in header:
                    self.worksheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + header['merge'] - 1)
                if 'style' in header:
                    cell.font = header['style'].get('font', Font())
                    cell.alignment = header['style'].get('alignment', Alignment())
                    cell.border = header['style'].get('border', Border())
                    cell.fill = header['style'].get('fill', PatternFill())
                col += header.get('merge', 1)
            row += 1

    def save(self):
        self.writer.close()

# Usage example
dataframe = pd.DataFrame({
    'A': [1, 2, 3],
    'B': [4, 5, 6],
    'C': [7, 8, 9],
    'D': [10, 11, 12]
})

header_config = [
    [{'label': 'DF Analysis', 'merge': 8, 'style': {'font': Font(bold=True, color='FFFFFF'), 'alignment': Alignment(horizontal='center'), 'fill': PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')}}],
    [{'label': 'DF1', 'merge': 4}, {'label': 'DF2', 'merge': 2}, {'label': 'DF Delta', 'merge': 2}],
    [{'label': 'A'}, {'label': 'B'}, {'label': 'C'}, {'label': 'D'}, {'label': 'C'}, {'label': 'D'}, {'label': 'C'}, {'label': 'D'}]
]

formatter = ExcelFormatter(dataframe, 7, 3, header_config)
formatter.write_data()
formatter.format_headers()
formatter.save()